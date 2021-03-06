import csv
from datetime import date
from openpyxl import load_workbook

wb = load_workbook(filename="in.xlsx")

total_sales = wb['Total Sales']
per_sales = wb['Percentage of Sales']
lookup = wb['Lookup Table']

lkp_dict = {k: v for v, k in lookup.iter_rows(min_row=2,values_only=True)}
total_dict = {(k, l.replace(" ", "").lower()): v for k,l,v in total_sales.iter_rows(min_row=2,values_only=True)}

with open('out.csv', 'w', newline='') as csvfile:
    writer = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Year Week Number', 'Scent', 'Size', 'Product Type', 'Sales'])
    for prod_id, date_in, size, prod_type, percent in per_sales.iter_rows(min_row=2,values_only=True):
        if percent != 0:
            year, week, _ = date(date_in.year,date_in.month,date_in.day).isocalendar()
            year_week = str(year) + '{:02d}'.format(week)
            name = lkp_dict[str(prod_id)+str(size)]
            sales = total_dict[(int(year_week), name.replace(" ", "").lower())] * percent
            writer.writerow([year_week,name,size,prod_type,'{:.2f}'.format(sales)])
